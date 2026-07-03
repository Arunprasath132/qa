from fastapi import APIRouter, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional, List
from groq import Groq
import os, io, json
import pandas as pd
from openpyxl import Workbook

router = APIRouter()

def get_client():
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY not set")
    return Groq(api_key=api_key)

def smart_rename(df):
    """Rename columns to standard names using lowercase + strip matching."""
    # First strip ALL column names of whitespace and special chars
    df.columns = [str(c).strip().replace('\xa0', ' ').replace('\t', ' ') for c in df.columns]

    print("DEBUG columns after strip:", list(df.columns))
    print("DEBUG columns repr:", [repr(c) for c in df.columns])

    rename = {}
    for col in df.columns:
        low = col.lower().strip()
        print(f"  checking: {repr(col)} → lower: {repr(low)}")

        if low in ["title", "test title", "test case title", "test name",
                   "name", "scenario", "description", "test case name"]:
            rename[col] = "title"

        elif low in ["steps", "test steps", "step", "steps to reproduce",
                     "execution steps", "actions", "procedure", "test procedure",
                     "test action", "test actions"]:
            rename[col] = "steps"

        elif low in ["expected", "expected result", "expected results",
                     "expected outcome", "expected output", "expected behavior",
                     "expected behaviour", "result", "outcome",
                     "pass criteria", "acceptance criteria", "expected value"]:
            rename[col] = "expected"

        elif low in ["id", "test id", "tc id", "test case id",
                     "sl no", "sl.no", "s.no", "sno", "no", "sr no", "sr.no"]:
            rename[col] = "id"

        elif low in ["type", "test type", "case type", "test case type"]:
            rename[col] = "type"

        elif low in ["priority", "test priority", "case priority"]:
            rename[col] = "priority"

    print("DEBUG rename map:", rename)
    df = df.rename(columns=rename)
    print("DEBUG columns after rename:", list(df.columns))
    return df


class TestCaseRequest(BaseModel):
    module: str
    feature: str
    user_story: Optional[str] = ""
    test_types: List[str] = ["positive", "negative", "validation", "boundary"]

SYSTEM_PROMPT = """You are an expert QA engineer. Generate test cases based on the provided information.
Return ONLY a valid JSON array with no markdown formatting, no code blocks, just raw JSON.
Each object must have: id, title, type, steps, expected, priority.
- id: sequential like TC_001, TC_002
- title: concise test case name
- type: positive/negative/validation/boundary
- steps: numbered steps as a single string separated by \\n
- expected: expected result
- priority: High/Medium/Low"""

@router.post("/generate")
async def generate_test_cases(request: TestCaseRequest):
    types_str = ", ".join(request.test_types)
    prompt = f"""Generate test cases for:
Module: {request.module}
Feature: {request.feature}
User Story: {request.user_story}
Test Types needed: {types_str}
Generate 3-5 test cases per type requested. Return as a JSON array only."""
    try:
        client = get_client()
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=2000
        )
        content = response.choices[0].message.content.strip()
        if content.startswith("```"):
            content = content.split("```")[1]
            if content.startswith("json"):
                content = content[4:]
        test_cases = json.loads(content.strip())
        return {"test_cases": test_cases, "count": len(test_cases)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export/excel")
async def export_excel(test_cases: List[dict]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    headers = ["ID", "Title", "Type", "Steps", "Expected Result", "Priority"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    for row, tc in enumerate(test_cases, 2):
        ws.cell(row=row, column=1, value=tc.get("id", ""))
        ws.cell(row=row, column=2, value=tc.get("title", ""))
        ws.cell(row=row, column=3, value=tc.get("type", ""))
        ws.cell(row=row, column=4, value=tc.get("steps", ""))
        ws.cell(row=row, column=5, value=tc.get("expected", ""))
        ws.cell(row=row, column=6, value=tc.get("priority", ""))
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"})


@router.post("/export/csv")
async def export_csv(test_cases: List[dict]):
    df = pd.DataFrame(test_cases)
    stream = io.StringIO()
    df.to_csv(stream, index=False)
    stream.seek(0)
    return StreamingResponse(iter([stream.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=test_cases.csv"})


@router.post("/upload")
async def upload_test_cases(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .csv files are supported")
    contents = await file.read()
    try:
        if file.filename.endswith(".xlsx"):
            # Read Excel — try multiple engines
            try:
                df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
            except Exception:
                df = pd.read_excel(io.BytesIO(contents))
        else:
            try:
                df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
            except Exception:
                df = pd.read_csv(io.StringIO(contents.decode("latin-1")))

        # Drop fully empty rows and columns
        df = df.dropna(how="all")
        df = df.dropna(axis=1, how="all")

        # Apply smart rename
        df = smart_rename(df)

        # Check required columns
        missing = {"title", "steps", "expected"} - set(df.columns)
        if missing:
            raise HTTPException(
                status_code=422,
                detail=f"Missing columns: {missing}. Found in file: {list(df.columns)}"
            )

        test_cases = df.to_dict(orient="records")
        for i, tc in enumerate(test_cases):
            # Clean NaN
            for key in list(tc.keys()):
                try:
                    if pd.isna(tc[key]):
                        tc[key] = ""
                except Exception:
                    pass
            # Ensure required fields are strings
            tc["id"]       = str(tc.get("id", f"TC_{str(i+1).zfill(3)}")).strip()
            tc["title"]    = str(tc.get("title", "")).strip()
            tc["steps"]    = str(tc.get("steps", "")).strip()
            tc["expected"] = str(tc.get("expected", "")).strip()
            tc.setdefault("type", "positive")
            tc.setdefault("priority", "Medium")

        return {"test_cases": test_cases, "count": len(test_cases)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"File parse error: {str(e)}")


@router.post("/debug-upload")
async def debug_upload(file: UploadFile = File(...)):
    """Debug: see exact column names Python reads from your file."""
    contents = await file.read()
    try:
        if file.filename.endswith(".xlsx"):
            df = pd.read_excel(io.BytesIO(contents), engine="openpyxl")
        else:
            df = pd.read_csv(io.StringIO(contents.decode("utf-8")))
        return {
            "columns_raw":   list(df.columns),
            "columns_lower": [str(c).lower().strip() for c in df.columns],
            "columns_repr":  [repr(c) for c in df.columns],
            "row_count":     len(df),
            "first_row":     df.iloc[0].to_dict() if len(df) > 0 else {}
        }
    except Exception as e:
        return {"error": str(e)}


@router.post("/from-screenshot")
async def generate_from_screenshot(file: UploadFile = File(...)):
    """Generate test cases from a UI screenshot using vision."""
    if not file.filename.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
        raise HTTPException(status_code=400, detail="Only image files are supported (PNG, JPG, JPEG, WEBP)")

    contents = await file.read()
    import base64

    # Encode image to base64
    b64_image = base64.b64encode(contents).decode("utf-8")
    ext = file.filename.split(".")[-1].lower()
    mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "webp": "image/webp"}.get(ext, "image/png")

    prompt = f"""You are an expert QA engineer. Analyze this UI screenshot and generate comprehensive test cases.

Look at the UI elements visible in the screenshot (buttons, forms, inputs, dropdowns, links, etc.)
and generate test cases covering:
- Positive test cases (happy path)
- Negative test cases (invalid inputs, errors)
- Validation test cases (field validations)
- Boundary test cases (edge cases)

Return ONLY a valid JSON array with no markdown, no code blocks.
Each object must have: id, title, type, steps, expected, priority.
- id: TC_001, TC_002, etc.
- title: concise test case name
- type: positive/negative/validation/boundary
- steps: numbered steps as single string separated by \\n
- expected: expected result
- priority: High/Medium/Low

Generate 8-12 test cases based on what you see in the screenshot."""

    try:
        client = get_client()

        # Use Groq vision model
        response = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:{mime};base64,{b64_image}"
                            }
                        },
                        {
                            "type": "text",
                            "text": prompt
                        }
                    ]
                }
            ],
            temperature=0.3,
            max_tokens=3000
        )

        content = response.choices[0].message.content.strip()
        if content.startswith("```"):
            content = content.split("```")[1]
            if content.startswith("json"):
                content = content[4:]
        test_cases = json.loads(content.strip())
        return {"test_cases": test_cases, "count": len(test_cases)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
